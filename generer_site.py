def generer_grille(fichier, nom_feuille):

    import openpyxl
    excel = openpyxl.load_workbook(fichier, data_only=True)
    feuille = excel[nom_feuille]

    padding_de_grille = '|&#8288 {: style="padding:0"}'

    grille = []

    # Écrire les entêtes du tableau de l'horaire
    entete = ""
    trait = ""
    for col in range(1, feuille.max_column + 1):
        entete += feuille.cell(row=1, column=col).value
        trait += "--"
        if col < feuille.max_column:
            entete += "|"
            trait += "|"
        else:
            entete += "\n"
            trait += "\n"

    grille.append(entete)
    grille.append(trait)

    #parcourir la grille pour toute les lignes
    for row in range(2, feuille.max_row + 1):
        ligne = ""
        has_colspan = False
        # pour toute les colonnes
        for col in range(1, feuille.max_column + 1):
            valeur = feuille.cell(row=row, column=col).value
            if valeur is None:
                valeur = ""
            else:
                valeur = str(valeur).replace("\n", "")

            #Petit bogue à l'affichage md si numéro semaine vide dans le excel
            if col == 1 and valeur == "":
                valeur ="|"

            #sanitaire
            valeur = valeur.strip()

            ligne += valeur

            if col < feuille.max_column:
                if not has_colspan:
                    ligne += "|"
            else:
                if has_colspan:
                    ligne += padding_de_grille * (feuille.max_column - 1)
                ligne += "\n"
            if "colspan" in valeur:
                has_colspan = True

    grille.append(ligne)
    return grille

    #fin generer_grille(fichier, nom_feuille)

def generer_horaire():

    grille = generer_grille("template/horaire.xlsx", "horaire")

    with open("./docs/horaire.md", "w") as f:
        # Écrire le titre de la page
        f.write("# Horaire du cours de bases de données 1\n")
        f.writelines(grille)

    #fin generer_horaire()



def main():
    print("Générer la page de l'horaire")
    generer_horaire()
    print ("fin de la génération")

    #fin main()

if __name__ == "__main__":
    main()

